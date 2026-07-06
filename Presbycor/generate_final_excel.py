import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import sys

csv_path = r'C:\Users\3D_OCT\Documents\Antigravity\Presbycor\Presbycor_Database_Completo.csv'
desktop = os.path.join(os.environ['USERPROFILE'], 'Desktop')
xlsx_path = os.path.join(desktop, 'Presbycor_Database_Final_Consolidado.xlsx')

print("Reading CSV...")
df = pd.read_csv(csv_path)

final_data = []

for idx, row in df.iterrows():
    # Helper to safely parse numeric
    def get_num(col):
        try: return float(row.get(col, 0))
        except: return 0.0

    od_dom = str(row.get("OD Dominant", "")).strip().lower() == "yes"
    os_dom = str(row.get("OS Dominant", "")).strip().lower() == "yes"
    od_add = get_num("OD Pre Min Add (D)")
    os_add = get_num("OS Pre Min Add (D)")
    
    # Selection logic
    chosen = ""
    reason = ""
    
    if od_dom and not os_dom:
        chosen = "OD"
        reason = "Dominant"
    elif os_dom and not od_dom:
        chosen = "OS"
        reason = "Dominant"
    else:
        # Tie-breaker: Highest Add
        if od_add >= os_add:
            chosen = "OD"
            reason = "Add"
        else:
            chosen = "OS"
            reason = "Add"
            
    # Extract features for chosen eye
    prefix = chosen + " "
    
    # We will build a unified dictionary without the eye prefix
    rec = {
        "Patient": row.get("Patient Name"),
        "Age": row.get("Age (years)"),
        "Chosen Eye": chosen,
        "Selection Reason": reason,
        
        # Pre-Op
        "Pre Sphere (D)": row.get(f"{prefix}Pre Sphere (D)"),
        "Pre Cylinder (D)": row.get(f"{prefix}Pre Cylinder (D)"),
        "Pre Axis (deg)": row.get(f"{prefix}Pre Axis (deg)"),
        "Pre K1 (D)": row.get(f"{prefix}Pre K1 (D)"),
        "Pre K2 (D)": row.get(f"{prefix}Pre K2 (D)"),
        "Pre Asph Q": row.get(f"{prefix}Pre Asph Q"),
        "Pre Min Add (D)": row.get(f"{prefix}Pre Min Add (D)"),
        "Pre Pachy (um)": row.get(f"{prefix}Pre Pachy (um)"),
        "Pre Pupillo (mm)": row.get(f"{prefix}Pre Pupillo (mm)"),
        
        # EQUI
        "EQUI Sphere (D)": row.get(f"{prefix}Equi Sphere (D)"),
        "EQUI Cylinder (D)": row.get(f"{prefix}Equi Cylinder (D)"),
        "EQUI Q Target": row.get(f"{prefix}Equi Q Target"),
        "EQUI OZ (mm)": row.get(f"{prefix}Equi OZ (mm)"),
        
        # DUAL
        "DUAL Sphere (D)": row.get(f"{prefix}Dual Sphere (D)"),
        "DUAL Cylinder (D)": row.get(f"{prefix}Dual Cylinder (D)"),
        "DUAL Q Target": row.get(f"{prefix}Dual Q Target"),
        "DUAL OZ (mm)": row.get(f"{prefix}Dual OZ (mm)"),
        
        # MONO
        "MONO Sphere (D)": row.get(f"{prefix}Mono Sphere (D)"),
        "MONO Cylinder (D)": row.get(f"{prefix}Mono Cylinder (D)"),
        "MONO Q Target": row.get(f"{prefix}Mono Q Target"),
        "MONO OZ (mm)": row.get(f"{prefix}Mono OZ (mm)"),
    }
    
    final_data.append(rec)

final_df = pd.DataFrame(final_data)

print("Generating Excel...")
wb = Workbook()
ws = wb.active
ws.title = 'Presbycor Database Final'

header_fills = {
    'Geral': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
    'Pre': PatternFill(start_color='9BBB59', end_color='9BBB59', fill_type='solid'),
    'EQUI': PatternFill(start_color='F79646', end_color='F79646', fill_type='solid'),
    'DUAL': PatternFill(start_color='8064A2', end_color='8064A2', fill_type='solid'),
    'MONO': PatternFill(start_color='4BACC6', end_color='4BACC6', fill_type='solid')
}

header_font = Font(bold=True, color='FFFFFF')
align_center = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

columns = list(final_df.columns)
ws.append(columns)

for col_idx, col_name in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.alignment = align_center
    cell.border = border
    
    if 'Pre' in col_name: cell.fill = header_fills['Pre']
    elif 'EQUI' in col_name: cell.fill = header_fills['EQUI']
    elif 'DUAL' in col_name: cell.fill = header_fills['DUAL']
    elif 'MONO' in col_name: cell.fill = header_fills['MONO']
    else: cell.fill = header_fills['Geral']

alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
green_fill = PatternFill(start_color='C4D79B', end_color='C4D79B', fill_type='solid') # Light green for dominant

for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=False), 2):
    # row is a list of values
    reason = row[3] # Index of 'Selection Reason'
    
    is_dominant = (reason == "Dominant")
    
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
        if is_dominant:
            cell.fill = green_fill
        elif r_idx % 2 == 0:
            cell.fill = alt_fill

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except: pass
    ws.column_dimensions[column].width = min(max_length + 2, 25)

ws.freeze_panes = 'B2'
ws.auto_filter.ref = ws.dimensions

wb.save(xlsx_path)
print(f"Excel gerado com sucesso em: {xlsx_path}")
