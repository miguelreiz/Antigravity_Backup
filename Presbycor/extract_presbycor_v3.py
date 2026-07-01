#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys, os
# Force unbuffered output for background task compatibility
os.environ['PYTHONUNBUFFERED'] = '1'

import fitz, easyocr, numpy as np, re, json, traceback, time, csv
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PRESBYCOR_DIR = Path(r"D:\Antigravity\Presbycor arquivos de pacientes")
DOWNLOADS_COMET_DIR = Path(r"D:\Downloads Comet")
OUTPUT_DIR = Path(r"C:\Users\3D_OCT\Documents\Antigravity\Presbycor")
OUTPUT_XLSX = OUTPUT_DIR / "Presbycor_Database_Completo.xlsx"
OUTPUT_CSV = OUTPUT_DIR / "Presbycor_Database_Completo.csv"
CHECKPOINT_DIR = OUTPUT_DIR / "checkpoints"
CHECKPOINT_FILE = CHECKPOINT_DIR / "checkpoint_v3.json"
CHECKPOINT_EVERY = 20
DPI = 200

READER = None  # Initialized in main()


# ─── HELPERS ─────────────────────────────────────────────────────────────────
def safe_float(v):
    if v is None: return None
    try: return float(str(v).strip().replace(",",".").replace(" ",""))
    except: return None

def extract_num(text):
    m = re.search(r"([+-]?\d+[.,]\d+)", str(text))
    if m: return m.group(1).replace(",",".")
    m = re.search(r"([+-]?\d+)", str(text))
    return m.group(1) if m else ""

def extract_pos_num(text):
    m = re.search(r"(\d+[.,]\d+)", str(text))
    if m: return m.group(1).replace(",",".")
    m = re.search(r"(\d{2,4})", str(text))
    return m.group(1) if m else ""

def extract_neg_num(text):
    m = re.search(r"([+-]\d+[.,]\d+)", str(text))
    if m: return m.group(1).replace(",",".")
    m = re.search(r"(\d+[.,]\d+)", str(text))
    return m.group(1).replace(",",".") if m else ""

def extract_int_only(text):
    m = re.search(r"\d+", str(text))
    return m.group() if m else ""

def extract_date(text):
    m = re.search(r"\d{2}/\d{2}/\d{4}", str(text))
    return m.group() if m else ""

# ─── PDF RENDER ───────────────────────────────────────────────────────────────
def render_pdf(path):
    doc = fitz.open(str(path))
    page = doc[0]
    mat = fitz.Matrix(DPI/72, DPI/72)
    pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
    img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, 3)
    doc.close()
    return img

# ─── OCR STRIP ────────────────────────────────────────────────────────────────
def ocr_strip(img, y1, y2, x1=0, x2=None):
    h, w = img.shape[:2]
    if x2 is None: x2 = w
    y1,y2 = max(0,y1),min(h,y2)
    x1,x2 = max(0,x1),min(w,x2)
    if y2<=y1 or x2<=x1: return []
    crop = img[y1:y2, x1:x2]
    if crop.size == 0: return []
    try: results = READER.readtext(crop, detail=1, paragraph=False)
    except: return []
    words = []
    for bbox, text, conf in results:
        if text.strip():
            xc = (bbox[0][0]+bbox[2][0])/2 + x1
            words.append((xc, str(text).strip()))
    return sorted(words, key=lambda t: t[0])

def find_after(words, label_re, max_gap=250):
    pat = re.compile(label_re, re.I)
    for i,(x,t) in enumerate(words):
        if pat.search(t):
            for j in range(i+1, len(words)):
                xj,tj = words[j]
                if xj-x >= 5 and xj-x <= max_gap: return tj
                elif xj-x > max_gap: break
    return ""

def extract_from_merged(token, pattern):
    """
    OCR often merges label+value into one token.
    e.g., "Sphere   +0.75" -> "+0.75"
    e.g., "K2   43.00 D" -> "43.00"
    e.g., "Target -0.82" -> "-0.82"
    e.g., "Asphericity Q -0.22" -> "-0.22"
    """
    # Try to find a number after the label pattern
    pat = re.compile(pattern, re.I)
    m = pat.search(token)
    if m:
        # Get text after the match
        rest = token[m.end():]
        # Extract signed decimal first
        nm = re.search(r"[+-]?\d+[.,]\d+", rest)
        if nm: return nm.group().replace(",",".")
        # Or plain decimal
        nm = re.search(r"\d+[.,]\d+", rest)
        if nm: return nm.group().replace(",",".")
        # Or integer
        nm = re.search(r"[+-]?\d+", rest)
        if nm: return nm.group()
    return ""

# ─── LAYOUT CONSTANTS (verified at 200 DPI) ──────────────────────────────────
# Coordinate system: y increases downward, x increases rightward
# Image size: 1653x2339 px

OD_X = (0, 826)    # Left half: OD
OS_X = (826, 1653)  # Right half: OS

# SECTION Y-OFFSETS (base=0 for PRE, value rows start at y_base + offset)
# Verified: PRE refraction y~1011, EQUI refraction y~1011-470=541 from top
# Wait - PRE is at y=520-560 (base), EQUI at y~990-1030 so offset = 990-520 = 470
SECTION_Y = {
    "pre":  0,
    "equi": 470,   # Was 490, corrected: refraction at y~990
    "dual": 810,   # Was 830, corrected: refraction at y~1330
    "mono": 1148,  # Was 1168, corrected: refraction at y~1668
}

# Q Target / Optical Zone absolute positions (verified)
OZ_ABS_Y = {
    "equi": 1157,  # y=1157-1202 works
    "dual": 1499,  # y=1499-1544 works
    "mono": 1838,  # y=1838-1883 works
}

# ─── SECTION PARSERS ──────────────────────────────────────────────────────────

def parse_refraction(img, y_off, x1, x2):
    """
    Parse 'Sphere +1.00  Cylinder +0.00  Axis 0°  VD 12 mm' row.
    Note: OCR sometimes merges "Sphere   +0.75" into one token.
    """
    y1, y2 = 520+y_off, 565+y_off
    w = ocr_strip(img, y1, y2, x1, x2)
    
    sphere = ""
    cylinder = ""
    axis = ""
    vd = ""
    
    for x, t in w:
        # Check for merged "Sphere +value" token
        if re.search(r"^spher", t, re.I):
            v = extract_from_merged(t, r"^spher")
            if v: sphere = v
            else: sphere = extract_num(find_after(w, r"^spher", 150))
        elif re.search(r"^cyl", t, re.I):
            v = extract_from_merged(t, r"^cyl")
            if v: cylinder = v
            else: cylinder = extract_num(find_after(w, r"^cyl", 150))
        # Handle merged "Axis   180" or "Axis 0"
        elif re.search(r"^axis", t, re.I):
            v = extract_from_merged(t, r"^axis")
            if v: axis = v
            else: axis = extract_int_only(find_after(w, r"^axis$", 80))
        elif re.search(r"^vd$", t, re.I):
            vd = extract_int_only(find_after(w, r"^vd$", 120))
        # merged "VD 12 mm"
        elif re.search(r"^vd\b", t, re.I):
            v = extract_from_merged(t, r"^vd\s*")
            if v: vd = v
    
    # Fallback if not found through iteration
    if not sphere: sphere = extract_num(find_after(w, r"^spher", 150))
    if not cylinder: cylinder = extract_num(find_after(w, r"^cyl", 150))
    if not axis:
        # Check any token with "Axis" in it
        for x, t in w:
            if re.search(r"axis", t, re.I):
                v = extract_from_merged(t, r"axis\s*")
                if v: axis = v; break
                else: axis = extract_int_only(find_after(w, r"^axis", 80))
    if not vd:
        for x, t in w:
            if re.search(r"vd", t, re.I):
                v = extract_from_merged(t, r"vd\s*")
                if v: vd = v; break
    
    return {"sphere": sphere, "cylinder": cylinder, "axis": axis, "vd": vd}

def parse_k_rows(img, y_off, x1, x2):
    """
    Parse K1 and K2 rows together (y=555-655). Wide strip to catch both rows.
    Wide scan reveals:
    - 'K2   43.00 D' at x~186 (merged, K2 row)
    - '43.00 D' at x~202 (K1 row overlapping)
    - 'Axis' at x~287, '180' at x~344 (K1 axis)
    - 'Pupillometry' at x~462, '3.0 mm' at x~556
    - 'Eccentricity E' at x~677, '0.47' at x~758
    - 'Asphericity Q -0.22' at x~703 (merged) or split
    
    Strategy: scan a wide strip, use x-position to disambiguate K1 vs K2
    """
    # Use tighter strips per row
    # K1 row at y~575-615 (but K2 label appears at wider strip)
    k1_w = ocr_strip(img, 570+y_off, 615+y_off, x1, x2)
    k2_w = ocr_strip(img, 605+y_off, 650+y_off, x1, x2)
    wide_w = ocr_strip(img, 555+y_off, 660+y_off, x1, x2)
    
    # K1: leftmost decimal in k1_w (no "K" label, just value at x~150-250)
    k1 = ""
    k1_axis = ""
    eccentricity_e = ""
    
    # Parse K1 from k1 strip: value at x~150-250 (local x, since we pass x1)
    for x, t in k1_w:
        lx = x - x1  # local x
        if 50 <= lx <= 300:
            v = extract_pos_num(t)
            if v and not k1: k1 = v
    # K1 axis: "Axis" then integer
    for x, t in k1_w:
        if re.search(r"^axis", t, re.I):
            v = extract_from_merged(t, r"^axis\s*")
            if v: k1_axis = v; break
    if not k1_axis:
        k1_axis = extract_int_only(find_after(k1_w, r"^axis$", 80))
    # Eccentricity E from k1 strip
    for x, t in k1_w:
        if re.search(r"eccent", t, re.I):
            v = extract_from_merged(t, r"eccent.*?[Ee]\s*")
            if v: eccentricity_e = v; break
    if not eccentricity_e:
        ecc_raw = find_after(k1_w, r"eccent", 150)
        if ecc_raw.lower() in ("e", "r", ""):
            ecc_raw = find_after(wide_w, r"eccent", 200)
        eccentricity_e = extract_pos_num(ecc_raw)
    
    # K2: parse from wide strip - look for "K2" merged token
    k2 = ""
    pupillometry = ""
    asphericity_q = ""
    
    for x, t in wide_w:
        lx = x - x1
        # K2 label+value merged token
        if re.search(r"^k2\b", t, re.I):
            v = extract_from_merged(t, r"^k2\s*")
            if v: k2 = v; break
    if not k2:
        # Fallback: look for value in K2 strip at x~150-250
        for x, t in k2_w:
            lx = x - x1
            if 50 <= lx <= 300:
                v = extract_pos_num(t)
                if v and not k2: k2 = v
    
    # Pupillometry from k2 strip
    for x, t in k2_w:
        if re.search(r"pupill", t, re.I):
            v = extract_from_merged(t, r"pupill.*?\s*")
            if v: pupillometry = v; break
    if not pupillometry:
        pupill_raw = find_after(k2_w, r"pupill", 150)
        if not pupill_raw: pupill_raw = find_after(wide_w, r"pupill", 150)
        pupillometry = extract_pos_num(pupill_raw)
    
    # Asphericity Q from k2 strip
    for x, t in k2_w:
        if re.search(r"aspher", t, re.I):
            v = extract_from_merged(t, r"aspher.*?[Qq]\s*")
            if v: asphericity_q = v; break
    if not asphericity_q:
        asph_raw = find_after(k2_w, r"aspher", 150)
        if asph_raw.lower() == "q":
            asph_raw = find_after(k2_w, r"^q$", 80)
        if not asph_raw: asph_raw = find_after(wide_w, r"aspher", 200)
        asphericity_q = extract_neg_num(asph_raw)
    
    return {
        "k1": k1, "k1_axis": k1_axis, "eccentricity_e": eccentricity_e,
        "k2": k2, "pupillometry": pupillometry, "asphericity_q": asphericity_q,
    }

def parse_other(img, y_off, x1, x2):
    """Parse min addition, CDVA, CNVA, pachymetry. Only in PRE section."""
    w = ocr_strip(img, 690+y_off, 730+y_off, x1, x2)
    return {
        "min_addition": extract_pos_num(find_after(w, r"addition|minimu", 150)),
        "cdva_log":     extract_pos_num(find_after(w, r"^cdva$", 100)),
        "cnva_log":     extract_pos_num(find_after(w, r"^cnva$", 100)),
        "pachymetry":   extract_int_only(find_after(w, r"pachym", 150)),
    }

def parse_oz(img, section, x1, x2):
    """
    Parse Optical Zone + Q Target row.
    EQUI: y=1157-1202; DUAL: y=1499-1544; MONO: y=1838-1883
    OCR may merge "Target -0.82" into one token.
    """
    base_y = OZ_ABS_Y.get(section, 1157)
    w = ocr_strip(img, base_y, base_y+45, x1, x2)
    
    optical_zone = ""
    q_target = ""
    
    for x, t in w:
        # Optical Zone value - look for decimal near x~190 (OD) or similar
        if re.search(r"optical|^oz$|^zone$", t, re.I):
            v = extract_from_merged(t, r"(optical|zone)\s*")
            if v: optical_zone = v
    
    if not optical_zone:
        optical_zone = extract_pos_num(find_after(w, r"optical|zone", 150))
    if not optical_zone:
        # Value at x~190 (6.50 mm)
        for x, t in w:
            lx = x - x1
            if 100 <= lx <= 280:
                v = extract_pos_num(t)
                if v: optical_zone = v; break
    
    # Q Target: look for "Target" label or merged "Target -0.82"
    for x, t in w:
        if re.search(r"target", t, re.I):
            v = extract_from_merged(t, r"target\s*")
            if v: q_target = v; break
    if not q_target:
        q_target = extract_neg_num(find_after(w, r"target", 120))
    
    return {"optical_zone": optical_zone, "q_target": q_target}

# ─── MAIN PDF PARSER ──────────────────────────────────────────────────────────

def parse_presbycor_pdf(pdf_path):
    img = render_pdf(pdf_path)
    h, w = img.shape[:2]
    rec = {"source_file": pdf_path.name, "source_folder": str(pdf_path.parent)}
    
    # HEADER
    # Values row: y=325-365
    # Verified positions: name x~166 (local x<420), DOB x~523, age x~930, surgeon x~1517
    hv = ocr_strip(img, 325, 365, 0, w)
    
    # Patient name: x<420 (verified: RAIMUNDO Silva at x=166)
    rec["patient_name"] = " ".join(t for x,t in hv if x<420).strip()
    # DOB: x between 420 and 750
    dob_txt = " ".join(t for x,t in hv if 420<=x<750)
    rec["dob"] = extract_date(dob_txt) or dob_txt.strip()
    # Age: x between 850 and 1100
    age_txt = " ".join(t for x,t in hv if 850<=x<1100)
    rec["age"] = extract_int_only(age_txt)
    # Surgeon: x >= 1300
    rec["surgeon"] = " ".join(t for x,t in hv if x>=1300).strip()
    # Creation date: top right y=95-125, x>=1100
    cw = ocr_strip(img, 95, 125, 1100, w)
    rec["creation_date"] = extract_date(" ".join(t for _,t in cw))
    
    # DOMINANT EYE (badge area y=435-478)
    od_d = ocr_strip(img, 435, 478, *OD_X)
    os_d = ocr_strip(img, 435, 478, *OS_X)
    com_o = " ".join(t for _,t in od_d).lower()
    com_s = " ".join(t for _,t in os_d).lower()
    rec["od_dominant"] = "yes" if "yes" in com_o else ("no" if "no" in com_o else "")
    rec["os_dominant"] = "yes" if "yes" in com_s else ("no" if "no" in com_s else "")
    
    # SECTIONS
    for sec, y_off in SECTION_Y.items():
        for eye, (x1, x2) in [("od", OD_X), ("os", OS_X)]:
            pfx = f"{eye}_{sec}"
            
            # Refraction
            ref = parse_refraction(img, y_off, x1, x2)
            for k,v in ref.items(): rec[f"{pfx}_{k}"] = v
            
            # K rows (K1, K2, Eccentricity, Pupillometry, Asphericity)
            kr = parse_k_rows(img, y_off, x1, x2)
            for k,v in kr.items(): rec[f"{pfx}_{k}"] = v
            
            if sec == "pre":
                # Other row (min addition, CDVA, CNVA, pachymetry)
                oth = parse_other(img, y_off, x1, x2)
                for k,v in oth.items(): rec[f"{pfx}_{k}"] = v
            else:
                # Optical Zone + Q Target
                oz = parse_oz(img, sec, x1, x2)
                for k,v in oz.items(): rec[f"{pfx}_{k}"] = v
    
    # ALERTS (y=755-920)
    aw = ocr_strip(img, 755, 920)
    rec["alerts"] = " ".join(t for _,t in aw)
    
    return rec

# ─── DEDUP & COLLECTION ───────────────────────────────────────────────────────

def get_base_name(f):
    return re.sub(r"\s*\(\d+\)\s*$", "", Path(f).stem).strip().lower()

def collect_all_pdfs():
    presbycor = list(PRESBYCOR_DIR.glob("**/*.pdf"))
    comet = list(DOWNLOADS_COMET_DIR.glob("strategy_*.pdf")) if DOWNLOADS_COMET_DIR.exists() else []
    in_p = {f.stem for f in presbycor if f.name.startswith("strategy_")}
    new_c = [f for f in comet if f.stem not in in_p]
    groups = {}
    for f in presbycor: groups.setdefault(get_base_name(f.name), []).append(f)
    selected = []; dupes = 0
    for base, grp in groups.items():
        no_suf = [f for f in grp if not re.search(r"\(\d+\)", f.name)]
        if no_suf:
            best = no_suf[0]
        else:
            def get_num(f):
                m = re.search(r"\((\d+)\)", f.name)
                return int(m.group(1)) if m else 0
            best = max(grp, key=get_num)
        selected.append(best)
        if len(grp)>1: dupes += len(grp)-1
    total = selected + new_c
    print(f"PDFs: {len(total)} ({dupes} dupes removed)")
    return total, dupes

# ─── CHECKPOINT ───────────────────────────────────────────────────────────────

def load_checkpoint():
    CHECKPOINT_DIR.mkdir(parents=True, exist_ok=True)
    if CHECKPOINT_FILE.exists():
        with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f: d = json.load(f)
        n_rec = len(d.get("records", []))
        print(f"Checkpoint: {n_rec} records done")
        return d
    return {"records":[], "processed":[], "errors":[]}

def save_checkpoint(d):
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f: json.dump(d, f, ensure_ascii=False, indent=2)
    n_rec = len(d.get("records", []))
    print(f"  [Checkpoint saved: {n_rec} records]")

# ─── EXCEL SCHEMA ─────────────────────────────────────────────────────────────

COLUMN_GROUPS = [
    ("Patient Info", "1F4E79", [
        ("source_file","Source File"),("patient_name","Patient Name"),("dob","Date of Birth"),
        ("age","Age (years)"),("creation_date","Creation Date"),("surgeon","Surgeon")]),
    ("OD - PRE OPERATIVE", "2E75B6", [
        ("od_dominant","OD Dominant"),
        ("od_pre_sphere","OD Pre Sphere (D)"),("od_pre_cylinder","OD Pre Cylinder (D)"),
        ("od_pre_axis","OD Pre Axis (deg)"),("od_pre_vd","OD Pre VD (mm)"),
        ("od_pre_k1","OD Pre K1 (D)"),("od_pre_k1_axis","OD Pre K1 Axis (deg)"),
        ("od_pre_k2","OD Pre K2 (D)"),("od_pre_eccentricity_e","OD Pre Ecc E"),
        ("od_pre_pupillometry","OD Pre Pupillo (mm)"),("od_pre_asphericity_q","OD Pre Asph Q"),
        ("od_pre_pachymetry","OD Pre Pachy (um)"),("od_pre_min_addition","OD Pre Min Add (D)"),
        ("od_pre_cdva_log","OD Pre CDVA Log"),("od_pre_cnva_log","OD Pre CNVA Log")]),
    ("OS - PRE OPERATIVE", "1A5EA8", [
        ("os_dominant","OS Dominant"),
        ("os_pre_sphere","OS Pre Sphere (D)"),("os_pre_cylinder","OS Pre Cylinder (D)"),
        ("os_pre_axis","OS Pre Axis (deg)"),("os_pre_vd","OS Pre VD (mm)"),
        ("os_pre_k1","OS Pre K1 (D)"),("os_pre_k1_axis","OS Pre K1 Axis (deg)"),
        ("os_pre_k2","OS Pre K2 (D)"),("os_pre_eccentricity_e","OS Pre Ecc E"),
        ("os_pre_pupillometry","OS Pre Pupillo (mm)"),("os_pre_asphericity_q","OS Pre Asph Q"),
        ("os_pre_pachymetry","OS Pre Pachy (um)"),("os_pre_min_addition","OS Pre Min Add (D)"),
        ("os_pre_cdva_log","OS Pre CDVA Log"),("os_pre_cnva_log","OS Pre CNVA Log")]),
    ("OD - EQUI-VISION", "00B0F0", [
        ("od_equi_sphere","OD Equi Sphere (D)"),("od_equi_cylinder","OD Equi Cylinder (D)"),
        ("od_equi_axis","OD Equi Axis (deg)"),("od_equi_vd","OD Equi VD (mm)"),
        ("od_equi_k1","OD Equi K1 (D)"),("od_equi_k1_axis","OD Equi K1 Axis (deg)"),
        ("od_equi_k2","OD Equi K2 (D)"),("od_equi_eccentricity_e","OD Equi Ecc E"),
        ("od_equi_pupillometry","OD Equi Pupillo (mm)"),("od_equi_asphericity_q","OD Equi Asph Q"),
        ("od_equi_optical_zone","OD Equi OZ (mm)"),("od_equi_q_target","OD Equi Q Target")]),
    ("OS - EQUI-VISION", "0088CC", [
        ("os_equi_sphere","OS Equi Sphere (D)"),("os_equi_cylinder","OS Equi Cylinder (D)"),
        ("os_equi_axis","OS Equi Axis (deg)"),("os_equi_vd","OS Equi VD (mm)"),
        ("os_equi_k1","OS Equi K1 (D)"),("os_equi_k1_axis","OS Equi K1 Axis (deg)"),
        ("os_equi_k2","OS Equi K2 (D)"),("os_equi_eccentricity_e","OS Equi Ecc E"),
        ("os_equi_pupillometry","OS Equi Pupillo (mm)"),("os_equi_asphericity_q","OS Equi Asph Q"),
        ("os_equi_optical_zone","OS Equi OZ (mm)"),("os_equi_q_target","OS Equi Q Target")]),
    ("OD - DUAL-VISION", "70AD47", [
        ("od_dual_sphere","OD Dual Sphere (D)"),("od_dual_cylinder","OD Dual Cylinder (D)"),
        ("od_dual_axis","OD Dual Axis (deg)"),("od_dual_vd","OD Dual VD (mm)"),
        ("od_dual_k1","OD Dual K1 (D)"),("od_dual_k1_axis","OD Dual K1 Axis (deg)"),
        ("od_dual_k2","OD Dual K2 (D)"),("od_dual_eccentricity_e","OD Dual Ecc E"),
        ("od_dual_pupillometry","OD Dual Pupillo (mm)"),("od_dual_asphericity_q","OD Dual Asph Q"),
        ("od_dual_optical_zone","OD Dual OZ (mm)"),("od_dual_q_target","OD Dual Q Target")]),
    ("OS - DUAL-VISION", "548235", [
        ("os_dual_sphere","OS Dual Sphere (D)"),("os_dual_cylinder","OS Dual Cylinder (D)"),
        ("os_dual_axis","OS Dual Axis (deg)"),("os_dual_vd","OS Dual VD (mm)"),
        ("os_dual_k1","OS Dual K1 (D)"),("os_dual_k1_axis","OS Dual K1 Axis (deg)"),
        ("os_dual_k2","OS Dual K2 (D)"),("os_dual_eccentricity_e","OS Dual Ecc E"),
        ("os_dual_pupillometry","OS Dual Pupillo (mm)"),("os_dual_asphericity_q","OS Dual Asph Q"),
        ("os_dual_optical_zone","OS Dual OZ (mm)"),("os_dual_q_target","OS Dual Q Target")]),
    ("OD - MONO-VISION", "FFC000", [
        ("od_mono_sphere","OD Mono Sphere (D)"),("od_mono_cylinder","OD Mono Cylinder (D)"),
        ("od_mono_axis","OD Mono Axis (deg)"),("od_mono_vd","OD Mono VD (mm)"),
        ("od_mono_k1","OD Mono K1 (D)"),("od_mono_k1_axis","OD Mono K1 Axis (deg)"),
        ("od_mono_k2","OD Mono K2 (D)"),("od_mono_eccentricity_e","OD Mono Ecc E"),
        ("od_mono_pupillometry","OD Mono Pupillo (mm)"),("od_mono_asphericity_q","OD Mono Asph Q"),
        ("od_mono_optical_zone","OD Mono OZ (mm)"),("od_mono_q_target","OD Mono Q Target")]),
    ("OS - MONO-VISION", "ED7D31", [
        ("os_mono_sphere","OS Mono Sphere (D)"),("os_mono_cylinder","OS Mono Cylinder (D)"),
        ("os_mono_axis","OS Mono Axis (deg)"),("os_mono_vd","OS Mono VD (mm)"),
        ("os_mono_k1","OS Mono K1 (D)"),("os_mono_k1_axis","OS Mono K1 Axis (deg)"),
        ("os_mono_k2","OS Mono K2 (D)"),("os_mono_eccentricity_e","OS Mono Ecc E"),
        ("os_mono_pupillometry","OS Mono Pupillo (mm)"),("os_mono_asphericity_q","OS Mono Asph Q"),
        ("os_mono_optical_zone","OS Mono OZ (mm)"),("os_mono_q_target","OS Mono Q Target")]),
    ("Alerts", "C00000", [("alerts","Alerts")]),
]

def get_all_cols():
    return [(fk,lbl,gn,c) for gn,c,fields in COLUMN_GROUPS for fk,lbl in fields]

def hex_fill(h): return PatternFill("solid", fgColor=h)
def mk_border(s="thin"):
    sd = Side(style=s, color="CCCCCC")
    return Border(left=sd, right=sd, top=sd, bottom=sd)
DARK = {"1F4E79":"163A5C","2E75B6":"1A5292","1A5EA8":"0F4A8A","00B0F0":"0090CC",
        "0088CC":"006699","70AD47":"538135","548235":"3D6226","FFC000":"CC9900",
        "ED7D31":"C45F10","C00000":"900000"}

def create_excel(records, out):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Presbycor DB"
    all_cols = get_all_cols()
    ws.row_dimensions[1].height = 28
    col = 1
    for gn, color, fields in COLUMN_GROUPS:
        s, e = col, col+len(fields)-1
        cell = ws.cell(row=1,column=s,value=gn)
        cell.fill = hex_fill(color)
        cell.font = Font(name="Calibri",bold=True,color="FFFFFF",size=10)
        cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        if e>s: ws.merge_cells(start_row=1,start_column=s,end_row=1,end_column=e)
        col += len(fields)
    ws.row_dimensions[2].height = 42
    for ci,(fk,lbl,gn,color) in enumerate(all_cols,1):
        c = ws.cell(row=2,column=ci,value=lbl)
        c.fill = hex_fill(DARK.get(color,color))
        c.font = Font(name="Calibri",bold=True,color="FFFFFF",size=9)
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border = mk_border("thin")
    fw=hex_fill("FFFFFF"); fb=hex_fill("EEF3FB")
    NUM={fk for fk,lbl,gn,c in all_cols if any(x in lbl for x in
         ["(D)","(mm)","(um)","(deg)","Log","years","OZ","Target"])}
    for ri,rec in enumerate(records,3):
        fill=fw if ri%2==1 else fb
        ws.row_dimensions[ri].height=15
        for ci,(fk,lbl,gn,color) in enumerate(all_cols,1):
            raw=rec.get(fk,"")
            val=safe_float(raw) if fk in NUM and raw and safe_float(raw) is not None else raw
            c=ws.cell(row=ri,column=ci,value=val)
            c.fill=fill; c.font=Font(name="Calibri",size=9)
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.border=mk_border("thin")
            if isinstance(val,float):
                c.number_format="0" if any(x in lbl for x in ["deg","um"]) else "+0.00;-0.00;0.00"
    for ci,(fk,lbl,gn,color) in enumerate(all_cols,1):
        cl=get_column_letter(ci)
        ws.column_dimensions[cl].width=(32 if "source_file" in fk else 22 if fk in
            ("patient_name","surgeon") else 55 if "alerts" in fk else 13 if
            "date" in fk or fk=="dob" else 9 if "dominant" in fk else 11)
    ws.freeze_panes="C3"
    ws.auto_filter.ref=f"A2:{get_column_letter(len(all_cols))}{len(records)+2}"
    # Stats sheet
    ws2=wb.create_sheet("Estatisticas")
    for ci,h in enumerate(["Campo","N","Media","DP","Min","Max","Mediana","% Preench"],1):
        c=ws2.cell(row=1,column=ci,value=h)
        c.fill=hex_fill("1F4E79"); c.font=Font(name="Calibri",bold=True,color="FFFFFF",size=10)
        c.alignment=Alignment(horizontal="center",vertical="center")
    ws2.column_dimensions["A"].width=38
    for ci in range(2,9): ws2.column_dimensions[get_column_letter(ci)].width=10
    total=len(records)
    stat_flds=[(fk,lbl) for fk,lbl,gn,c in all_cols if fk in NUM]
    for ri,(fk,lbl) in enumerate(stat_flds,2):
        vals=[v for v in (safe_float(r.get(fk,"")) for r in records) if v is not None]
        n=len(vals)
        if n:
            mean=sum(vals)/n; std=(sum((x-mean)**2 for x in vals)/n)**0.5
            mn,mx=min(vals),max(vals); sv=sorted(vals); mid=n//2
            med=sv[mid] if n%2==1 else (sv[mid-1]+sv[mid])/2
        else: mean=std=mn=mx=med=None
        pct=round(n/total*100,1) if total else 0
        fill2=fw if ri%2==1 else fb
        row2=[lbl,n,round(mean,3) if mean is not None else "",
              round(std,3) if std is not None else "",
              round(mn,3) if mn is not None else "",
              round(mx,3) if mx is not None else "",
              round(med,3) if med is not None else "",pct]
        for ci,val in enumerate(row2,1):
            c2=ws2.cell(row=ri,column=ci,value=val)
            c2.fill=fill2; c2.font=Font(name="Calibri",size=9)
            c2.alignment=Alignment(horizontal="center",vertical="center")
            c2.border=mk_border("thin")
    ws3=wb.create_sheet("Resumo")
    ws3.column_dimensions["A"].width=30; ws3.column_dimensions["B"].width=65
    ws3.cell(row=1,column=1,value="Presbycor Database v3").font=Font(name="Calibri",bold=True,size=14,color="1F4E79")
    for ri,(k,v) in enumerate([
        ("Data:",datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Registros:",total),("Colunas:",len(all_cols)),
        ("Pasta:",str(PRESBYCOR_DIR)),("Excel:",str(out))],3):
        ws3.cell(row=ri,column=1,value=k).font=Font(name="Calibri",bold=True,size=10)
        ws3.cell(row=ri,column=2,value=v).font=Font(name="Calibri",size=10)
    wb.save(str(out))
    print(f"Excel saved: {out}")

def create_csv(records, out):
    all_cols=get_all_cols(); fks=[fk for fk,lbl,gn,c in all_cols]
    hdrs={fk:lbl for fk,lbl,gn,c in all_cols}
    with open(str(out),"w",newline="",encoding="utf-8-sig") as f:
        w=csv.DictWriter(f,fieldnames=fks,extrasaction="ignore")
        w.writerow(hdrs)
        for rec in records: w.writerow({fk:rec.get(fk,"") for fk in fks})
    print(f"CSV saved: {out}")

def coverage_report(records):
    all_cols=get_all_cols(); total=len(records)
    if not total: return
    print(f"\n=== COVERAGE ({total} records) ===")
    for fk,lbl,gn,c in all_cols:
        n=sum(1 for r in records if str(r.get(fk,"")).strip()!="")
        pct=n/total*100
        bar="#"*int(pct/5)+"."*(20-int(pct/5))
        print(f"  {lbl[:34]:<34} {bar} {pct:5.1f}%")

def main():
    global READER
    print("="*70, flush=True)
    print("PRESBYCOR PDF EXTRACTOR v3 - Calibrated Strip OCR", flush=True)
    print(f"Started: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", flush=True)
    print("="*70, flush=True)
    print("Initializing EasyOCR...", flush=True)
    READER = easyocr.Reader(["en"], gpu=False, verbose=False)
    print("EasyOCR Ready!", flush=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    all_files,_ = collect_all_pdfs()
    ckpt = load_checkpoint()
    done = set(ckpt["processed"]); records=ckpt["records"]; errors=ckpt["errors"]
    to_do = [f for f in all_files if str(f) not in done]
    print(f"To process: {len(to_do)} | Done: {len(done)}\n", flush=True)

    t0=time.time(); times=[]
    for i,pdf_path in enumerate(to_do):
        t_file=time.time()
        eta_str=f"ETA: {sum(times)/len(times)*(len(to_do)-i)/60.0:.1f}min" if times else "ETA: ..."
        print(f"[{i+1}/{len(to_do)}] {pdf_path.name[:60]} | {eta_str}", flush=True)
        try:
            rec=parse_presbycor_pdf(pdf_path)
            if rec:
                records.append(rec)
                elapsed=time.time()-t_file; times.append(elapsed)
                print(f"  [OK] {rec.get('patient_name','N/A')} | {rec.get('creation_date','N/A')} | {elapsed:.1f}s", flush=True)
            else: errors.append({"file":str(pdf_path),"error":"Empty"})
        except Exception as e:
            errors.append({"file":str(pdf_path),"error":str(e)})
            print(f"  [ERR] {e}", flush=True); traceback.print_exc()
        ckpt["processed"].append(str(pdf_path))
        if (i+1)%CHECKPOINT_EVERY==0:
            ckpt["records"]=records; ckpt["errors"]=errors; save_checkpoint(ckpt)
    ckpt["records"]=records; ckpt["errors"]=errors; save_checkpoint(ckpt)
    print(f"\nDone: {len(records)} records | {len(errors)} errors | {(time.time()-t0)/60.0:.1f}min", flush=True)
    if not records: return
    coverage_report(records)
    print(f"\nGenerating Excel..."); create_excel(records, OUTPUT_XLSX)
    print("Generating CSV..."); create_csv(records, OUTPUT_CSV)
    if errors:
        print(f"\nERRORS ({len(errors)}):")
        for e in errors: print(f"  {Path(e['file']).name}: {e['error']}")
    print(f"\nDone: {OUTPUT_XLSX}")

if __name__=="__main__":
    main()
