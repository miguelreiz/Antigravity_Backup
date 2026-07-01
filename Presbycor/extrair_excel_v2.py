import sys, os, re, glob, json, base64, urllib.request, subprocess
sys.stdout.reconfigure(encoding="utf-8")

PDF_DIRS = [
    r"D:\Antigravity\Presbycor arquivos de pacientes",
    r"D:\Downloads Comet",
]
OUT_EXCEL = r"C:\Users\3D_OCT\Documents\Antigravity\Presbycor\Presbycor_Database_Completo.xlsx"
OUT_CSV   = r"C:\Users\3D_OCT\Documents\Antigravity\Presbycor\Presbycor_Database_Completo.csv"

def pip(pkg):
    subprocess.run([sys.executable, "-m", "pip", "install", pkg, "-q"], check=False)

try: import fitz
except: pip("pymupdf"); import fitz
try: import pandas as pd
except: pip("pandas"); import pandas as pd
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except: pip("openpyxl")
try: import numpy as np
except: pip("numpy"); import numpy as np

# Gemini key via subprocess (usando o proprio python do antigravity)
import ctypes, os
GEMINI_KEY = os.environ.get("GEMINI_API_KEY","")
if not GEMINI_KEY:
    # Tentar ler do registry ou outros locais
    try:
        import winreg
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Environment")
        GEMINI_KEY, _ = winreg.QueryValueEx(key, "GEMINI_API_KEY")
    except: pass

print(f"API Key: {'OK ('+str(len(GEMINI_KEY))+' chars)' if GEMINI_KEY else 'NAO ENCONTRADA'}")

def pdf_to_png(path, dpi=250):
    doc = fitz.open(path)
    pix = doc[0].get_pixmap(matrix=fitz.Matrix(dpi/72, dpi/72), alpha=False)
    return pix.tobytes("png")

PROMPT = """Extraia todos os dados deste relatorio Presbycor Treatment Planning.
O relatorio tem colunas OD (esquerda) e OS (direita).
Para cada olho extrai: PRE OPERATIVE (Sphere, Cylinder, Axis, K1 D, K2 D, Eccentricity E, Asphericity Q, Pupillometry mm, Pachymetry um, Minimum addition D, CDVA Log, CNVA Log).
Para cada estrategia (EQUI-VISION, DUAL-VISION, MONO-VISION OPTIMIZED): Sphere, Cylinder, Axis, K1, K2, Eccentricity E, Asphericity Q, Pupillometry, Optical Zone mm, Q Target.
Tambem: Patient name, Date of birth, Age, Creation date, Surgeon, Dominant eye (OD or OS).
Retorne JSON com esta estrutura exata (use null para valores nao encontrados):
{
  "paciente": "", "data_nasc": "", "idade": null, "data_calculo": "", "cirurgiao": "",
  "olho_dominante": "",
  "OD": {
    "pre_esfera": null, "pre_cilindro": null, "pre_eixo": null,
    "K1": null, "K2": null, "excentricidade": null, "Q_pre": null,
    "pupilometria": null, "paquimetria": null, "adicao_min": null,
    "CDVA": null, "CNVA": null,
    "equi_esfera": null, "equi_cilindro": null, "equi_eixo": null, "equi_ZO": null, "equi_Q_target": null,
    "dual_esfera": null, "dual_cilindro": null, "dual_eixo": null, "dual_ZO": null, "dual_Q_target": null,
    "mono_esfera": null, "mono_cilindro": null, "mono_eixo": null, "mono_ZO": null, "mono_Q_target": null
  },
  "OS": {
    "pre_esfera": null, "pre_cilindro": null, "pre_eixo": null,
    "K1": null, "K2": null, "excentricidade": null, "Q_pre": null,
    "pupilometria": null, "paquimetria": null, "adicao_min": null,
    "CDVA": null, "CNVA": null,
    "equi_esfera": null, "equi_cilindro": null, "equi_eixo": null, "equi_ZO": null, "equi_Q_target": null,
    "dual_esfera": null, "dual_cilindro": null, "dual_eixo": null, "dual_ZO": null, "dual_Q_target": null,
    "mono_esfera": null, "mono_cilindro": null, "mono_eixo": null, "mono_ZO": null, "mono_Q_target": null
  },
  "alertas_OD": [], "alertas_OS": [], "notas": ""
}
Retorne APENAS JSON valido."""

def ocr_gemini(img_bytes, nome, key):
    if not key: return {}
    payload = {
        "contents": [{"parts": [
            {"text": PROMPT},
            {"inline_data": {"mime_type": "image/png", "data": base64.b64encode(img_bytes).decode()}}
        ]}],
        "generationConfig": {"temperature": 0, "maxOutputTokens": 3000}
    }
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={key}"
    try:
        req = urllib.request.Request(url, data=json.dumps(payload).encode(),
                                     headers={"Content-Type": "application/json"}, method="POST")
        with urllib.request.urlopen(req, timeout=30) as r:
            res = json.loads(r.read())
        txt = res["candidates"][0]["content"]["parts"][0]["text"]
        txt = re.sub(r"```json\s*|```", "", txt).strip()
        return json.loads(txt)
    except Exception as e:
        print(f"  Gemini erro: {e}")
        return {}

# Coletar PDFs unicos
pdfs_all = []
for d in PDF_DIRS:
    if os.path.exists(d):
        pdfs_all += glob.glob(os.path.join(d, "**", "*.pdf"), recursive=True)

seen = {}
for p in pdfs_all:
    key_n = re.sub(r"\s*\(\d+\)\.pdf$", ".pdf", os.path.basename(p), flags=re.I).lower()
    if key_n not in seen: seen[key_n] = p
pdfs = list(seen.values())

# Filtrar apenas PDFs Presbycor (strategy_XXXX ou Presbycor_*)
pdfs_presby = [p for p in pdfs if re.search(r"(strategy_\d+|presbycor_)", os.path.basename(p), re.I)]
print(f"PDFs Presbycor unicos: {len(pdfs_presby)}")

registros = []
for i, pdf in enumerate(pdfs_presby):
    nome = os.path.basename(pdf)
    id_strat = re.search(r"strategy_(\d+)", nome, re.I)
    id_strat = id_strat.group(1) if id_strat else ""
    print(f"[{i+1:3d}/{len(pdfs_presby)}] {nome[:55]}")

    rec = {"N": i+1, "ID": id_strat, "arquivo": nome}

    try:
        img = pdf_to_png(pdf)
        dados = ocr_gemini(img, nome, GEMINI_KEY)
        if dados:
            rec.update({
                "Paciente":      dados.get("paciente", nome.replace(".pdf","")),
                "Data Nasc":     dados.get("data_nasc",""),
                "Idade":         dados.get("idade"),
                "Data Calculo":  dados.get("data_calculo",""),
                "Cirurgiao":     dados.get("cirurgiao",""),
                "Dominante":     dados.get("olho_dominante",""),
                # OD Pre-op
                "OD Esf Pre":    dados.get("OD",{}).get("pre_esfera"),
                "OD Cil Pre":    dados.get("OD",{}).get("pre_cilindro"),
                "OD Eixo Pre":   dados.get("OD",{}).get("pre_eixo"),
                "OD K1":         dados.get("OD",{}).get("K1"),
                "OD K2":         dados.get("OD",{}).get("K2"),
                "OD Exc E":      dados.get("OD",{}).get("excentricidade"),
                "OD Q Pre":      dados.get("OD",{}).get("Q_pre"),
                "OD Pupila":     dados.get("OD",{}).get("pupilometria"),
                "OD Pachy":      dados.get("OD",{}).get("paquimetria"),
                "OD Add Min":    dados.get("OD",{}).get("adicao_min"),
                "OD CDVA":       dados.get("OD",{}).get("CDVA"),
                "OD CNVA":       dados.get("OD",{}).get("CNVA"),
                # OD Equi
                "OD Equi Esf":   dados.get("OD",{}).get("equi_esfera"),
                "OD Equi Cil":   dados.get("OD",{}).get("equi_cilindro"),
                "OD Equi ZO":    dados.get("OD",{}).get("equi_ZO"),
                "OD Equi Q":     dados.get("OD",{}).get("equi_Q_target"),
                # OD Dual
                "OD Dual Esf":   dados.get("OD",{}).get("dual_esfera"),
                "OD Dual Cil":   dados.get("OD",{}).get("dual_cilindro"),
                "OD Dual ZO":    dados.get("OD",{}).get("dual_ZO"),
                "OD Dual Q":     dados.get("OD",{}).get("dual_Q_target"),
                # OD Mono
                "OD Mono Esf":   dados.get("OD",{}).get("mono_esfera"),
                "OD Mono Cil":   dados.get("OD",{}).get("mono_cilindro"),
                "OD Mono ZO":    dados.get("OD",{}).get("mono_ZO"),
                "OD Mono Q":     dados.get("OD",{}).get("mono_Q_target"),
                # OS Pre-op
                "OS Esf Pre":    dados.get("OS",{}).get("pre_esfera"),
                "OS Cil Pre":    dados.get("OS",{}).get("pre_cilindro"),
                "OS Eixo Pre":   dados.get("OS",{}).get("pre_eixo"),
                "OS K1":         dados.get("OS",{}).get("K1"),
                "OS K2":         dados.get("OS",{}).get("K2"),
                "OS Exc E":      dados.get("OS",{}).get("excentricidade"),
                "OS Q Pre":      dados.get("OS",{}).get("Q_pre"),
                "OS Pupila":     dados.get("OS",{}).get("pupilometria"),
                "OS Pachy":      dados.get("OS",{}).get("paquimetria"),
                "OS Add Min":    dados.get("OS",{}).get("adicao_min"),
                "OS CDVA":       dados.get("OS",{}).get("CDVA"),
                "OS CNVA":       dados.get("OS",{}).get("CNVA"),
                # OS Equi
                "OS Equi Esf":   dados.get("OS",{}).get("equi_esfera"),
                "OS Equi Cil":   dados.get("OS",{}).get("equi_cilindro"),
                "OS Equi ZO":    dados.get("OS",{}).get("equi_ZO"),
                "OS Equi Q":     dados.get("OS",{}).get("equi_Q_target"),
                # OS Dual
                "OS Dual Esf":   dados.get("OS",{}).get("dual_esfera"),
                "OS Dual Cil":   dados.get("OS",{}).get("dual_cilindro"),
                "OS Dual ZO":    dados.get("OS",{}).get("dual_ZO"),
                "OS Dual Q":     dados.get("OS",{}).get("dual_Q_target"),
                # OS Mono
                "OS Mono Esf":   dados.get("OS",{}).get("mono_esfera"),
                "OS Mono Cil":   dados.get("OS",{}).get("mono_cilindro"),
                "OS Mono ZO":    dados.get("OS",{}).get("mono_ZO"),
                "OS Mono Q":     dados.get("OS",{}).get("mono_Q_target"),
                # Alertas
                "Alertas OD":    "; ".join(dados.get("alertas_OD",[])),
                "Alertas OS":    "; ".join(dados.get("alertas_OS",[])),
                "Notas":         dados.get("notas",""),
            })
        else:
            rec["Paciente"] = re.sub(r"strategy_\d+", "", nome).replace(".pdf","").replace("Presbycor_\d{2}-\d{2}-\d{4}_","").replace("_"," ").strip()
    except Exception as e:
        rec["Paciente"] = nome.replace(".pdf","")
        rec["Notas"] = f"ERRO: {e}"

    registros.append(rec)
    # Salvar incrementalmente a cada 10 casos
    if (i+1) % 10 == 0:
        pd.DataFrame(registros).to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
        print(f"  Checkpoint: {i+1} casos salvos")

df = pd.DataFrame(registros)
df.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
print(f"\nCSV salvo: {OUT_CSV} ({len(df)} casos, {len(df.columns)} colunas)")

# Verificar cobertura
print("\nCobertura dos campos principais:")
campos_check = ["Paciente","Idade","OD K1","OD Q Pre","OD Equi Esf","OD Equi Q","OD Dual Esf","OS K1","OS Equi Esf"]
for c in campos_check:
    if c in df.columns:
        n = df[c].notna().sum()
        print(f"  {c:20s}: {n}/{len(df)} ({100*n/len(df):.0f}%)")

# Gerar Excel profissional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Banco Presbycor"

thin = Side(style="thin", color="D0D0D0")
brd = Border(left=thin, right=thin, top=thin, bottom=thin)

GRUPOS = [
    ("IDENTIFICACAO",          "1B3A6B", ["N","ID","Paciente","Data Nasc","Idade","Data Calculo","Cirurgiao","Dominante"]),
    ("OD — PRE-OPERATORIO",    "7B2D00", ["OD Esf Pre","OD Cil Pre","OD Eixo Pre","OD K1","OD K2","OD Exc E","OD Q Pre","OD Pupila","OD Pachy","OD Add Min","OD CDVA","OD CNVA"]),
    ("OD — EQUI-VISION",       "1A5276", ["OD Equi Esf","OD Equi Cil","OD Equi ZO","OD Equi Q"]),
    ("OD — DUAL-VISION",       "117A65", ["OD Dual Esf","OD Dual Cil","OD Dual ZO","OD Dual Q"]),
    ("OD — MONO-VISION",       "5C1F8A", ["OD Mono Esf","OD Mono Cil","OD Mono ZO","OD Mono Q"]),
    ("OS — PRE-OPERATORIO",    "922B21", ["OS Esf Pre","OS Cil Pre","OS Eixo Pre","OS K1","OS K2","OS Exc E","OS Q Pre","OS Pupila","OS Pachy","OS Add Min","OS CDVA","OS CNVA"]),
    ("OS — EQUI-VISION",       "1F618D", ["OS Equi Esf","OS Equi Cil","OS Equi ZO","OS Equi Q"]),
    ("OS — DUAL-VISION",       "0E6655", ["OS Dual Esf","OS Dual Cil","OS Dual ZO","OS Dual Q"]),
    ("OS — MONO-VISION",       "76448A", ["OS Mono Esf","OS Mono Cil","OS Mono ZO","OS Mono Q"]),
    ("ALERTAS / NOTAS",        "515A5A", ["Alertas OD","Alertas OS","Notas","arquivo"]),
]

col_order = []
col_offset = 1
for grp, cor, cols in GRUPOS:
    cols_ok = [c for c in cols if c in df.columns]
    if not cols_ok: continue
    ws.merge_cells(start_row=1, start_column=col_offset, end_row=1, end_column=col_offset+len(cols_ok)-1)
    c = ws.cell(1, col_offset, grp)
    c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    c.fill = PatternFill("solid", fgColor=cor)
    c.alignment = Alignment(horizontal="center", vertical="center")
    for j, campo in enumerate(cols_ok):
        c2 = ws.cell(2, col_offset+j, campo)
        c2.font = Font(bold=True, color="FFFFFF", size=8, name="Calibri")
        c2.fill = PatternFill("solid", fgColor=cor)
        c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col_order.append(campo)
    col_offset += len(cols_ok)

for ri, (_, row) in enumerate(df.iterrows()):
    bg = "F2F3F4" if ri%2==0 else "FFFFFF"
    for ci, campo in enumerate(col_order):
        val = row.get(campo)
        if isinstance(val, float) and pd.isna(val): val = None
        c = ws.cell(ri+3, ci+1, val)
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = brd
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(size=9, name="Calibri")
        if "Esf" in campo or "Cil" in campo:
            c.number_format = "+0.00;-0.00;0.00"
        elif "K1" in campo or "K2" in campo:
            c.number_format = "0.00"

larguras = {"N":5,"ID":10,"Paciente":28,"Data Nasc":12,"Idade":7,"Data Calculo":12,
            "Cirurgiao":20,"Dominante":8,"Alertas OD":35,"Alertas OS":35,"Notas":40,"arquivo":35}
for ci, campo in enumerate(col_order):
    ws.column_dimensions[get_column_letter(ci+1)].width = larguras.get(campo, 11)

ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 30
ws.freeze_panes = "D3"

# Aba estatisticas
ws2 = wb.create_sheet("Estatisticas")
ws2.cell(1,1,"PRESBYCOR - Estatisticas Descritivas (N=" + str(len(df)) + ")").font = Font(bold=True,size=13)
num_cols = [c for c in col_order if c not in ["N","ID","Paciente","Data Nasc","Data Calculo","Cirurgiao","Dominante","Alertas OD","Alertas OS","Notas","arquivo"]]
for j,h in enumerate(["Variavel","N","Media","DP","Min","P25","Mediana","P75","Max"]):
    c = ws2.cell(4,j+1,h)
    c.font = Font(bold=True,color="FFFFFF")
    c.fill = PatternFill("solid",fgColor="1B3A6B")
    c.alignment = Alignment(horizontal="center")
r = 5
for col in num_cols:
    if col not in df.columns: continue
    s = pd.to_numeric(df[col], errors="coerce").dropna()
    if len(s) < 2: continue
    for j,v in enumerate([col, len(s), round(s.mean(),2), round(s.std(),2), round(s.min(),2), round(s.quantile(.25),2), round(s.median(),2), round(s.quantile(.75),2), round(s.max(),2)]):
        c = ws2.cell(r,j+1,v)
        c.fill = PatternFill("solid",fgColor="F2F3F4" if r%2==0 else "FFFFFF")
        c.border = brd
        c.alignment = Alignment(horizontal="center")
        c.font = Font(size=9)
    r += 1

wb.save(OUT_EXCEL)
print(f"\nExcel salvo: {OUT_EXCEL}")
print(f"Total final: {len(df)} casos | {len(df.columns)} colunas")
