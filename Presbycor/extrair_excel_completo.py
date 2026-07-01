#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PRESBYCOR — Extrator Completo para Excel Profissional
Estilo tabelas Renato Ambrosio — todos os campos clinicos
PDFs sao imagem pura, usa Gemini Vision para OCR
"""
import sys, os, re, glob, subprocess, json, base64, urllib.request
sys.stdout.reconfigure(encoding='utf-8')

PDF_DIRS = [
    r"D:\Antigravity\Presbycor arquivos de pacientes",
    r"D:\Downloads Comet",
]
OUT_EXCEL = r"C:\Users\3D_OCT\Documents\Antigravity\Presbycor\Presbycor_Database_Completo.xlsx"
OUT_CSV   = r"C:\Users\3D_OCT\Documents\Antigravity\Presbycor\Presbycor_Database_Completo.csv"

def pip(pkg):
    subprocess.run([sys.executable, "-m", "pip", "install", pkg, "-q"], check=False)

print("Verificando dependencias...")
try: import fitz
except: pip("pymupdf"); import fitz
try: import pandas as pd
except: pip("pandas"); import pandas as pd
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except: pip("openpyxl"); import openpyxl; from openpyxl.styles import PatternFill,Font,Alignment,Border,Side; from openpyxl.utils import get_column_letter
print("OK\n")

# Carregar API key
GEMINI_KEY = None
cfg_path = r"C:\Users\3D_OCT\.gemini\config\config.json"
if os.path.exists(cfg_path):
    with open(cfg_path) as f:
        cfg = json.load(f)
    GEMINI_KEY = cfg.get("apiKey") or cfg.get("api_key")
print(f"Gemini API: {'OK' if GEMINI_KEY else 'NAO ENCONTRADA'}")

def pdf_to_png(path, dpi=200):
    doc = fitz.open(path)
    pix = doc[0].get_pixmap(matrix=fitz.Matrix(dpi/72, dpi/72), alpha=False)
    return pix.tobytes("png")

def ocr_gemini(img_bytes, nome):
    if not GEMINI_KEY: return {}
    prompt = """Extraia TODOS os dados deste relatorio Presbycor/PresbyLASIK e retorne JSON:
{"nome_paciente":"","data":"","idade":null,"sexo":"","estrategia":"","olho_dominante":"",
 "OD":{"esfera":null,"cilindro":null,"eixo":null,"AV_sc_longe":null,"AV_cc_longe":null,"AV_perto":null,
  "K1":null,"K2":null,"Kmed":null,"Kmax":null,"Q_pre":null,"Q_alvo":null,
  "zona_optica":null,"pupila_escotp":null,"paquimetria":null,"RSB":null,"PTA":null,
  "esfera_target":null,"cilindro_target":null,"adicao_D":null,"ablacao_central_um":null,"ablacao_total_um":null,"flap_um":null},
 "OE":{"esfera":null,"cilindro":null,"eixo":null,"AV_sc_longe":null,"AV_cc_longe":null,"AV_perto":null,
  "K1":null,"K2":null,"Kmed":null,"Kmax":null,"Q_pre":null,"Q_alvo":null,
  "zona_optica":null,"pupila_escotp":null,"paquimetria":null,"RSB":null,"PTA":null,
  "esfera_target":null,"cilindro_target":null,"adicao_D":null,"ablacao_central_um":null,"ablacao_total_um":null,"flap_um":null},
 "notas":""}
Retorne APENAS o JSON valido, sem texto extra."""
    payload = {"contents":[{"parts":[{"text":prompt},{"inline_data":{"mime_type":"image/png","data":base64.b64encode(img_bytes).decode()}}]}],"generationConfig":{"temperature":0,"maxOutputTokens":2048}}
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_KEY}"
    try:
        req = urllib.request.Request(url, data=json.dumps(payload).encode(), headers={"Content-Type":"application/json"}, method="POST")
        with urllib.request.urlopen(req, timeout=45) as r:
            res = json.loads(r.read())
        txt = res["candidates"][0]["content"]["parts"][0]["text"]
        txt = re.sub(r"`json\s*|`","", txt).strip()
        return json.loads(txt)
    except Exception as e:
        print(f"    ERRO Gemini {nome}: {e}")
        return {}

def extrair_nome(path):
    n = os.path.basename(path).replace(".pdf","")
    m = re.search(r"Presbycor_[\d-]+_(.+)", n, re.I)
    if m: return m.group(1).replace("_"," ").title()
    m = re.search(r"strategy_(\d+)", n, re.I)
    if m: return f"Caso #{m.group(1)}"
    return n

# Coletar e deduplicar PDFs
pdfs_raw = []
for d in PDF_DIRS:
    if os.path.exists(d):
        pdfs_raw += glob.glob(os.path.join(d,"**","*.pdf"), recursive=True)

seen = {}
for p in pdfs_raw:
    key = re.sub(r"\s*\(\d+\)\.pdf$",".pdf", os.path.basename(p), flags=re.I).lower()
    if key not in seen: seen[key] = p
pdfs = list(seen.values())
print(f"PDFs unicos: {len(pdfs)} (de {len(pdfs_raw)} encontrados)\n")

# Extrair
registros = []
for i, pdf in enumerate(pdfs):
    nome_arq = os.path.basename(pdf)
    print(f"[{i+1:3d}/{len(pdfs)}] {nome_arq[:55]}")
    rec = {"N": i+1, "arquivo": nome_arq, "id": re.search(r"strategy_(\d+)", nome_arq) and re.search(r"strategy_(\d+)", nome_arq).group(1) or "", "nome_paciente": extrair_nome(pdf)}
    try:
        img = pdf_to_png(pdf)
        dados = ocr_gemini(img, nome_arq)
        if dados:
            rec["nome_paciente"] = dados.get("nome_paciente") or rec["nome_paciente"]
            rec["data"] = dados.get("data","")
            rec["idade"] = dados.get("idade")
            rec["sexo"] = dados.get("sexo","")
            rec["estrategia"] = dados.get("estrategia","")
            rec["olho_dominante"] = dados.get("olho_dominante","")
            for olho in ["OD","OE"]:
                o = dados.get(olho,{})
                for campo in ["esfera","cilindro","eixo","AV_sc_longe","AV_cc_longe","AV_perto","K1","K2","Kmed","Kmax","Q_pre","Q_alvo","zona_optica","pupila_escotp","paquimetria","RSB","PTA","esfera_target","cilindro_target","adicao_D","ablacao_central_um","ablacao_total_um","flap_um"]:
                    rec[f"{olho}_{campo}"] = o.get(campo)
            rec["notas"] = dados.get("notas","")
    except Exception as e:
        rec["notas"] = f"ERRO: {e}"
        print(f"    ERRO: {e}")
    registros.append(rec)

df = pd.DataFrame(registros)
df.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
print(f"\nCSV: {OUT_CSV}")

# Excel Profissional
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Banco Presbycor"

thin = Side(style="thin", color="D0D0D0")
brd = Border(left=thin,right=thin,top=thin,bottom=thin)

GRUPOS = [
    ("IDENTIFICACAO", "1B3A6B", ["N","id","nome_paciente","data","idade","sexo"]),
    ("ESTRATEGIA",    "2C6E3F", ["estrategia","olho_dominante"]),
    ("OD PRE-OP",     "7B2D00", ["OD_esfera","OD_cilindro","OD_eixo","OD_AV_sc_longe","OD_AV_cc_longe","OD_AV_perto"]),
    ("OD TOPOGRAFIA", "5C1F8A", ["OD_K1","OD_K2","OD_Kmed","OD_Kmax","OD_Q_pre"]),
    ("OD CIRURGIA",   "1A5276", ["OD_zona_optica","OD_pupila_escotp","OD_paquimetria","OD_RSB","OD_PTA","OD_flap_um"]),
    ("OD TARGETS",    "117A65", ["OD_esfera_target","OD_cilindro_target","OD_adicao_D","OD_Q_alvo","OD_ablacao_central_um","OD_ablacao_total_um"]),
    ("OE PRE-OP",     "7D3C98", ["OE_esfera","OE_cilindro","OE_eixo","OE_AV_sc_longe","OE_AV_cc_longe","OE_AV_perto"]),
    ("OE TOPOGRAFIA", "884EA0", ["OE_K1","OE_K2","OE_Kmed","OE_Kmax","OE_Q_pre"]),
    ("OE CIRURGIA",   "1F618D", ["OE_zona_optica","OE_pupila_escotp","OE_paquimetria","OE_RSB","OE_PTA","OE_flap_um"]),
    ("OE TARGETS",    "0E6655", ["OE_esfera_target","OE_cilindro_target","OE_adicao_D","OE_Q_alvo","OE_ablacao_central_um","OE_ablacao_total_um"]),
    ("NOTAS",         "515A5A", ["notas","arquivo"]),
]

LABELS = {
    "N":"N","id":"ID Estrategia","nome_paciente":"Paciente","data":"Data","idade":"Idade","sexo":"Sexo",
    "estrategia":"Estrategia","olho_dominante":"Dom.",
    "OD_esfera":"Esfera (D)","OD_cilindro":"Cilindro (D)","OD_eixo":"Eixo","OD_AV_sc_longe":"AVsc L","OD_AV_cc_longe":"AVcc L","OD_AV_perto":"AV P",
    "OD_K1":"K1 (D)","OD_K2":"K2 (D)","OD_Kmed":"Kmed","OD_Kmax":"Kmax","OD_Q_pre":"Q pre",
    "OD_zona_optica":"ZO (mm)","OD_pupila_escotp":"Pupila","OD_paquimetria":"Pachy","OD_RSB":"RSB","OD_PTA":"PTA%","OD_flap_um":"Flap (um)",
    "OD_esfera_target":"Esf.Target","OD_cilindro_target":"Cil.Target","OD_adicao_D":"Adicao (D)","OD_Q_alvo":"Q alvo","OD_ablacao_central_um":"Abl.C (um)","OD_ablacao_total_um":"Abl.T (um)",
    "OE_esfera":"Esfera (D)","OE_cilindro":"Cilindro (D)","OE_eixo":"Eixo","OE_AV_sc_longe":"AVsc L","OE_AV_cc_longe":"AVcc L","OE_AV_perto":"AV P",
    "OE_K1":"K1 (D)","OE_K2":"K2 (D)","OE_Kmed":"Kmed","OE_Kmax":"Kmax","OE_Q_pre":"Q pre",
    "OE_zona_optica":"ZO (mm)","OE_pupila_escotp":"Pupila","OE_paquimetria":"Pachy","OE_RSB":"RSB","OE_PTA":"PTA%","OE_flap_um":"Flap (um)",
    "OE_esfera_target":"Esf.Target","OE_cilindro_target":"Cil.Target","OE_adicao_D":"Adicao (D)","OE_Q_alvo":"Q alvo","OE_ablacao_central_um":"Abl.C (um)","OE_ablacao_total_um":"Abl.T (um)",
    "notas":"Notas","arquivo":"Arquivo PDF",
}

col_order = []
col_offset = 1
for grp_name, grp_color, cols in GRUPOS:
    cols_ok = [c for c in cols if c in df.columns]
    n = len(cols_ok)
    if n == 0: continue
    ws.merge_cells(start_row=1, start_column=col_offset, end_row=1, end_column=col_offset+n-1)
    c = ws.cell(1, col_offset, grp_name)
    c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    c.fill = PatternFill("solid", fgColor=grp_color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    for j, campo in enumerate(cols_ok):
        c2 = ws.cell(2, col_offset+j, LABELS.get(campo, campo))
        c2.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        c2.fill = PatternFill("solid", fgColor=grp_color)
        c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col_order.append(campo)
    col_offset += n

for ri, (_, row) in enumerate(df.iterrows()):
    bg = "F2F3F4" if ri%2==0 else "FDFEFE"
    for ci, campo in enumerate(col_order):
        val = row.get(campo)
        if isinstance(val, float) and pd.isna(val): val = None
        c = ws.cell(ri+3, ci+1, val)
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = brd
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(size=9, name="Calibri")
        if campo in ["OD_esfera","OD_cilindro","OD_esfera_target","OD_cilindro_target",
                     "OE_esfera","OE_cilindro","OE_esfera_target","OE_cilindro_target"]:
            c.number_format = "+0.00;-0.00;0.00"

for ci, campo in enumerate(col_order):
    w = {"nome_paciente":28,"notas":35,"arquivo":38,"estrategia":14}.get(campo, 10)
    ws.column_dimensions[get_column_letter(ci+1)].width = w

ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 28
ws.freeze_panes = "D3"

# Aba de estatisticas
ws2 = wb.create_sheet("Estatisticas")
ws2.cell(1,1,"PRESBYTERCOR - Estatisticas Descritivas").font = Font(bold=True,size=13)
ws2.cell(2,1,f"N = {len(df)} casos unicos")
num_cols = [c for c in col_order if c not in ["N","id","nome_paciente","data","sexo","estrategia","olho_dominante","notas","arquivo"]]
hdrs = ["Variavel","N","Media","DP","Min","P25","Mediana","P75","Max"]
for j,h in enumerate(hdrs):
    c = ws2.cell(4,j+1,h)
    c.font = Font(bold=True,color="FFFFFF")
    c.fill = PatternFill("solid",fgColor="1B3A6B")
    c.alignment = Alignment(horizontal="center")
r = 5
for col in num_cols:
    if col not in df.columns: continue
    s = pd.to_numeric(df[col],errors="coerce").dropna()
    if len(s)<2: continue
    for j,v in enumerate([LABELS.get(col,col),len(s),round(s.mean(),2),round(s.std(),2),round(s.min(),2),round(s.quantile(.25),2),round(s.median(),2),round(s.quantile(.75),2),round(s.max(),2)]):
        c = ws2.cell(r,j+1,v)
        c.fill = PatternFill("solid",fgColor="F2F3F4" if r%2==0 else "FDFEFE")
        c.border = brd
        c.alignment = Alignment(horizontal="center")
        c.font = Font(size=9)
    r+=1

wb.save(OUT_EXCEL)
print(f"\nExcel salvo: {OUT_EXCEL}")
print(f"Total: {len(df)} casos")
